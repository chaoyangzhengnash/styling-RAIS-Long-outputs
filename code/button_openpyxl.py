import openpyxl
import tkinter as tk
import os
import timeit
from openpyxl.styles import Font
from tkinter.filedialog import askopenfilename, asksaveasfilename, askdirectory

# Module function 
list_tables_name = list()
list_tables_path = list()

# create series of formatting style
bold             = Font(bold=True)
shalow_red       = Font(color='ff0000')
deep_red         = Font(color='bd0000')
shalow_red_bold  = Font(bold=True, color='ff0000')
deep_red_bold    = Font(bold=True, color='bd0000')
yellow_fill      = openpyxl.styles.PatternFill(start_color='FFFD4D', end_color='FFFD4D', fill_type='solid')
red_fill         = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def get_macro(ws):
    """Create marco variables(list) to reference indicators(by index) and flags"""
    ind_names   = list()
    ind_release = list()
    ind_status  = list()
    ind_symbol  = list()
    cols = [cell.value for cell in ws[1]]
    for i, col in enumerate(cols):
        if col.endswith('_release'): 
            ind_release.append(i)
            ind_names.append(next(i for i, j in enumerate(cols) if j == col[:-8]))
        if col.endswith('_status'): 
            ind_status.append(i)
        if col.endswith('_symbol'): 
            ind_symbol.append(i)    
    return [ind_names,ind_symbol,ind_status,ind_release]    

def formating(ws,lst,limit):
    """ Formating indicators based on flags """
    for row in ws.iter_rows(min_row=2):    
        # Highlight series with cohot_size exceed user defiend limit and not None
        if row[lst[0][0]].value is not None: #cohort size always the 1st indicator 
            if int(row[lst[0][0]].value) > int(limit):
                row[0].fill = red_fill # formating "year"                     
        # Bolding indicators based on symbol flags       
        for i, flag in enumerate(lst[1]):
            if row[flag].value == 1: # preliminary
                row[lst[0][i]].font = bold                  
        # Coloring status flags accordingly
        for i, flag in enumerate(lst[2]):
            if row[flag].value == 6: # Data quality: acceptable
                if row[lst[1][i]].value == 1:
                    row[lst[0][i]].font = shalow_red_bold 
                else:
                    row[lst[0][i]].font = shalow_red                
            if row[flag].value == 7: # Data quality: caution
                if row[lst[1][i]].value == 1:
                    row[lst[0][i]].font = deep_red_bold 
                else:
                    row[lst[0][i]].font = deep_red                    
        #Highlight indicators with release flags == 1
        for i, flag in enumerate(lst[3]):
            if row[flag].value == 1: 
                row[lst[0][i]].fill = yellow_fill 
    return ws
        
def forward(lst_name, lst_path, limit):
    """Go through the whole process."""
    start = timeit.default_timer()
    for i in lst_path:
        # Get output path
        filepath = asksaveasfilename(
                defaultextension="xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],)        
        # Processing input data 
        wb = openpyxl.load_workbook(i)
        ws = wb[wb.sheetnames[0]] # focus on the sheet "processed" 
        ls = get_macro(ws)
        formating(ws,ls,limit)
        wb.save(filepath)            
        wb.close()
    stop = timeit.default_timer()
    print('Time: ', stop - start)     

# Button function
def openSingleXlsx(lbl):
    """Open a sigle xlsx for editing."""
    table_path = askopenfilename(
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    list_tables_name.append(os.path.split(table_path)[1])
    list_tables_path.append(table_path)
    lbl["text"] = f"{str(list_tables_name)}"
    return list_tables_name, list_tables_path
    
def openFolder(lbl):
    """Open a folder and read all CODR tables for editing."""
    folder_path = askdirectory(title='Select Folder') 
    for root, dirs, filenames in os.walk(folder_path):
        for file in filenames:
            if not file.endswith("_MetaData.csv"):
                list_tables_name.append(file)
                list_tables_path.append(os.path.join(root, file))
                lbl["text"] = f"{str(list_tables_name)}"
    return list_tables_name, list_tables_path

def processing(lbl,limit):
    """Button to start the processing status."""
    lbl["text"] = "Start processing!"
    lbl.update()    
    forward(list_tables_name, list_tables_path, limit)    
    lbl["text"] = "Task done!"


    