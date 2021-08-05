""" Examine the saved excel file
    Highlight the indicators if it's going to be released
    The required structure of the excel file is that the excel file is in the
    'wide' form;
        - the very first row has column names
        - for each indicator, there is a release
            e.g., for an indicator <ind>, there's another column <ind_release>
                and the <ind> is going to be released if <ind_release> = 1.
"""

import openpyxl

""" path and file names of the excel file
        - in_path: the path of the original (unprocessed) file
        - original_file: the file name, including the extension (.xlsx)
        - out_path: the path of the output (processed) file
        - processed_file: the file name, including the extension (.xlsx)
"""
in_path = '\\\\Ctces01\\post2\\STUDENTS - RAIS\\6000_RecordLinkage\\Longitudinal project\\Longitudinal Indicators\\output\\'
original_file = 'pathway.xlsx'
out_path = '\\\\Ctces01\\post2\\STUDENTS - RAIS\\6000_RecordLinkage\\Longitudinal project\\Longitudinal Indicators\\python\\highlight release\\'
#out_path = in_path # same as the in_path
processed_file = 'pathway_highlighted.xlsx'

# open the excel file
wb = openpyxl.load_workbook(in_path+original_file)
# open the first sheet of the file
ws = wb[wb.sheetnames[0]]

# reads in the first row: assuming that it contains column names
cols = [cell.value for cell in ws[1]]

""" from the column names, find where the release flags are
    and the location of corresponding indicators
    assuming the flag name ends with _release with the corresponding indicator
    name in front of the suffix.
    e.g., <indicator>_release is the release flag for indicator <indicator>
"""
release_index = []
var_index = []

for i, col in enumerate(cols):
    if col.endswith('_release'): # if the column name ends with _release
        release_index.append(i) # save its index
        var_index.append(next(i for i, j in enumerate(cols) if j == col[:-8])) # and save the index of corresponding indicator
        
# create formatting style
red_font = openpyxl.styles.Font(bold=True, color='9C0103')
red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# iterate over the all rows
for row in ws.iter_rows(min_row=2):
    for i, flag in enumerate(release_index):
        if row[flag].value == 1: # if the release flag has value 1
            row[var_index[i]].fill = red_fill # apply the format on the indicator
            
wb.save(out_path+processed_file)            
wb.close()
